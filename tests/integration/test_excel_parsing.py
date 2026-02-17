"""Tests for Excel parsing logic (ai-panda-team.xlsx)."""

from pathlib import Path

import pytest

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def parse_team_from_excel(filepath):
    """Parse team members from Excel file.

    Expected column layout:
    A=naam, B=functie, C=team, D=telefoon (skipped), E=foto_url, F=email

    Returns list of dicts with keys: naam, functie, team, foto_url, email
    """
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        members.append({
            "naam": str(row[0] or "").strip(),
            "functie": str(row[1] or "").strip() if len(row) > 1 else "",
            "team": str(row[2] or "").strip() if len(row) > 2 else "",
            # column D (index 3) = telefoon, skipped
            "foto_url": str(row[4] or "").strip() if len(row) > 4 else "",
            "email": str(row[5] or "").strip() if len(row) > 5 else "",
        })
    wb.close()
    return members


@pytest.mark.integration
@pytest.mark.skipif(not HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelParsing:
    """Test Excel parsing logic with in-memory workbooks."""

    @pytest.fixture
    def create_test_workbook(self, tmp_path):
        """Create a test Excel file with known data."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # Header row
        ws.append(["Naam", "Functie", "Team", "Telefoon", "Foto URL", "Email"])
        # Data rows
        ws.append(["Alice", "Developer", "Alpha", "0612345678", "https://example.com/alice.jpg", "alice@test.nl"])
        ws.append(["Bob", "Designer", "Beta", "0687654321", "https://example.com/bob.jpg", "bob@test.nl"])
        ws.append(["Charlie", "Manager", "Alpha", None, None, "charlie@test.nl"])

        path = tmp_path / "test_team.xlsx"
        wb.save(path)
        return path

    def test_basic_column_mapping(self, create_test_workbook):
        """naam, functie, team, foto_url, email should be correctly mapped."""
        members = parse_team_from_excel(create_test_workbook)
        assert len(members) == 3
        alice = members[0]
        assert alice["naam"] == "Alice"
        assert alice["functie"] == "Developer"
        assert alice["team"] == "Alpha"
        assert alice["foto_url"] == "https://example.com/alice.jpg"
        assert alice["email"] == "alice@test.nl"

    def test_telefoon_column_skipped(self, create_test_workbook):
        """Column D (telefoon) should not appear in the output."""
        members = parse_team_from_excel(create_test_workbook)
        for m in members:
            assert "telefoon" not in m
            assert "0612345678" not in m.values()

    def test_empty_name_rows_filtered(self, tmp_path):
        """Rows with empty name should be skipped."""
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Naam", "Functie", "Team", "Telefoon", "Foto URL", "Email"])
        ws.append(["Alice", "Dev", "A", "06", "url", "a@b.nl"])
        ws.append([None, "Empty", "B", "06", "url", "e@b.nl"])
        ws.append(["", "Also Empty", "C", "06", "url", "f@b.nl"])
        ws.append(["Bob", "Design", "D", "06", "url", "b@b.nl"])

        path = tmp_path / "test_empty.xlsx"
        wb.save(path)
        members = parse_team_from_excel(path)
        assert len(members) == 2
        assert members[0]["naam"] == "Alice"
        assert members[1]["naam"] == "Bob"

    def test_none_values_fallback_to_empty_string(self, create_test_workbook):
        """None values in cells should become empty strings."""
        members = parse_team_from_excel(create_test_workbook)
        charlie = members[2]
        assert charlie["foto_url"] == "None" or charlie["foto_url"] == ""
        # Note: str(None) = "None", so we test the actual behavior
        # The parsing uses str(row[4] or ""), so None -> ""
        assert charlie["foto_url"] == ""

    def test_real_excel_file(self, project_root):
        """Parse the actual ai-panda-team.xlsx if it exists."""
        excel_path = project_root / "data" / "ai-panda-team.xlsx"
        if not excel_path.exists():
            pytest.skip("ai-panda-team.xlsx not found in data/")

        members = parse_team_from_excel(excel_path)
        assert len(members) >= 1, "Expected at least 1 team member"
        for m in members:
            assert m["naam"], "Every member should have a name"
