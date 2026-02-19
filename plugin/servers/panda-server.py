#!/usr/bin/env python3
"""
MCP server voor AI Panda plugin.
- AI-beeldgeneratie via Google Gemini (primair) en OpenAI (fallback)
- Panda referentie-image support (multimodal Gemini)
- Bedrijfslogo ophalen en meegeven
- Image upload naar catbox.moe / tmpfiles.org
- Team Excel uitlezen
"""

import asyncio
import base64
import json
import os
import subprocess
import sys
import tempfile
import urllib.request
from pathlib import Path

# Ensure dependencies
try:
    from mcp.server.fastmcp import FastMCP
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "mcp[cli]", "--break-system-packages", "-q"])
    from mcp.server.fastmcp import FastMCP

try:
    import httpx
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "httpx", "httpcore[socks]", "--break-system-packages", "-q"])
    import httpx

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl


mcp = FastMCP("panda-server")

# --- API Keys ---
# Loaded from environment (.mcp.json passes them through).
# Fallback: .env in project root (local development only).
if not os.environ.get("GEMINI_API_KEY") or not os.environ.get("OPENAI_API_KEY"):
    try:
        from dotenv import load_dotenv
        _env = Path(__file__).parent.parent.parent / ".env"
        if _env.exists():
            load_dotenv(_env)
    except ImportError:
        pass

_gemini_api_key = os.environ.get("GEMINI_API_KEY", "")
_openai_api_key = os.environ.get("OPENAI_API_KEY", "")
_logo_dev_token = os.environ.get("LOGO_DEV_TOKEN", "")

# --- Gemini config ---
GEMINI_MODEL = "gemini-3-pro-image-preview"
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

# --- Panda reference image ---
_panda_reference_b64: str | None = None

def _load_panda_reference() -> str | None:
    """Load panda-reference.png as base64 from known locations."""
    candidates = []
    plugin_root = os.environ.get("CLAUDE_PLUGIN_ROOT", "")
    if plugin_root:
        candidates.append(Path(plugin_root) / "assets" / "panda-reference.png")
    candidates.append(Path(__file__).parent.parent / "assets" / "panda-reference.png")
    candidates.append(Path(__file__).parent.parent.parent / "assets" / "panda-reference.png")

    for path in candidates:
        if path.exists():
            return base64.b64encode(path.read_bytes()).decode()
    return None

_panda_reference_b64 = _load_panda_reference()

# --- Panda character prompt (photorealistic, based on reference) ---
PANDA_PROMPT_BASE = (
    "A photorealistic giant panda with a realistic furry head, black eye patches, "
    "and round ears, on a human body wearing a tailored black business suit, "
    "crisp white dress shirt, and a bold orange necktie. The panda has black "
    "furry paws instead of human hands. Confident posture, striding forward as "
    "a leader. Cinematic photography style, natural lighting, sharp focus, "
    "shallow depth of field."
)

LOGO_PROMPT_WITH_REF = (
    "The company logo is provided as a reference image. Reproduce this logo accurately: "
    "on a large screen or digital display in the background, "
    "as a small badge or pin on the panda's suit lapel, "
    "and on any whiteboards, documents, or signage visible in the scene. "
    "Keep the logo's colors, proportions, and text exactly as shown."
)


def _build_panda_prompt(bedrijfsnaam: str, sector: str = "", has_logo: bool = False) -> str:
    """Build the full panda image prompt with sector-specific setting."""
    parts = [PANDA_PROMPT_BASE]

    if sector:
        parts.append(f"Setting: a modern {sector} environment related to {bedrijfsnaam}. "
                      f"The panda is presenting AI solutions to the team. "
                      f"Colleagues in appropriate professional attire walk behind.")
    else:
        parts.append(f"Setting: a modern corporate office. "
                      f"The panda is presenting AI solutions on a large screen. "
                      f"Colleagues in smart business attire walk behind.")

    if has_logo:
        parts.append(LOGO_PROMPT_WITH_REF)
    else:
        parts.append(f'A screen/whiteboard shows "{bedrijfsnaam}" in a professional corporate font.')

    return " ".join(parts)


# --- Logo fetching ---

def _fetch_logo_b64(domain: str) -> str | None:
    """Fetch company logo as base64. Tries Logo.dev (high quality), then Google Favicons.
    Uses curl subprocess to avoid Python SSL certificate issues."""
    if not domain:
        return None
    # Strip protocol if present
    domain = domain.replace("https://", "").replace("http://", "").strip("/")

    urls = []
    # Logo.dev: high quality logos, requires LOGO_DEV_TOKEN
    if _logo_dev_token:
        urls.append(f"https://img.logo.dev/{domain}?token={_logo_dev_token}&format=png&size=256")
    # Google Favicons: free, no key, lower quality (favicons only)
    urls.append(f"https://t0.gstatic.com/faviconV2?client=SOCIAL&type=FAVICON&fallback_opts=TYPE,SIZE,URL&url=http://{domain}&size=128")

    for url in urls:
        try:
            result = subprocess.run(
                ["curl", "-sL", "-o", "-", url],
                capture_output=True, timeout=10,
            )
            if result.returncode == 0 and len(result.stdout) > 100:
                return base64.b64encode(result.stdout).decode()
        except Exception:
            continue
    return None


# --- API Key tools ---

@mcp.tool()
async def check_api_keys() -> str:
    """
    Check welke API keys beschikbaar zijn in deze sessie.
    Retourneert JSON: {"gemini": true/false, "openai": true/false}
    """
    return json.dumps({
        "gemini": bool(_gemini_api_key),
        "openai": bool(_openai_api_key),
    })


@mcp.tool()
async def set_api_key(provider: str, api_key: str) -> str:
    """
    Sla een API key op in het geheugen van de MCP server (voor de duur van de sessie).
    De key wordt NIET naar schijf geschreven en verdwijnt wanneer de sessie eindigt.

    Args:
        provider: "gemini" of "openai"
        api_key: De API key.
    """
    global _gemini_api_key, _openai_api_key
    if not api_key or not api_key.strip():
        return json.dumps({"success": False, "error": "Lege key opgegeven."})

    provider = provider.strip().lower()
    key = api_key.strip()

    if provider == "gemini":
        _gemini_api_key = key
    elif provider == "openai":
        _openai_api_key = key
    else:
        return json.dumps({"success": False, "error": f"Onbekende provider: {provider}. Gebruik 'gemini' of 'openai'."})

    return json.dumps({"success": True, "provider": provider, "message": "API key opgeslagen voor deze sessie (alleen in geheugen)."})


# --- Image generation ---

async def generate_with_gemini(prompt: str, logo_b64: str | None = None) -> bytes | None:
    """Generate an image using Google Gemini API (multimodal with reference images)."""
    if not _gemini_api_key:
        return None

    # Build multimodal parts
    prompt_text = prompt
    if _panda_reference_b64 or logo_b64:
        prompt_text = "Generate an image matching the style of the reference photo. " + prompt

    parts: list[dict] = [{"text": prompt_text}]
    if _panda_reference_b64:
        parts.append({"inlineData": {"mimeType": "image/png", "data": _panda_reference_b64}})
    if logo_b64:
        parts.append({"inlineData": {"mimeType": "image/png", "data": logo_b64}})

    payload = {
        "contents": [{"parts": parts}],
        "generationConfig": {
            "responseModalities": ["TEXT", "IMAGE"],
        },
    }

    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(
            GEMINI_URL,
            params={"key": _gemini_api_key},
            json=payload,
        )
        resp.raise_for_status()
        data = resp.json()

    for candidate in data.get("candidates", []):
        for part in candidate.get("content", {}).get("parts", []):
            if "inlineData" in part:
                return base64.b64decode(part["inlineData"]["data"])
    return None


async def generate_with_openai(prompt: str) -> bytes | None:
    """Generate an image using OpenAI API (prompt-only, no reference images)."""
    if not _openai_api_key:
        return None

    payload = {
        "model": "gpt-image-1.5",
        "prompt": prompt,
        "n": 1,
        "size": "1024x1024",
        "quality": "high",
    }

    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(
            "https://api.openai.com/v1/images/generations",
            headers={
                "Authorization": f"Bearer {_openai_api_key}",
                "Content-Type": "application/json",
            },
            json=payload,
        )
        resp.raise_for_status()
        data = resp.json()

    # OpenAI returns b64_json or url; try b64 first, then download url
    for item in data.get("data", []):
        if "b64_json" in item:
            return base64.b64decode(item["b64_json"])
        if "url" in item:
            try:
                img_resp = urllib.request.urlopen(item["url"], timeout=30)
                return img_resp.read()
            except Exception:
                pass
    return None


# --- Upload helpers ---

async def upload_to_catbox(tmp_path: str) -> str:
    """Upload a file to catbox.moe and return the public URL."""
    result = subprocess.run(
        ["curl", "-s", "-F", "reqtype=fileupload", "-F", f"fileToUpload=@{tmp_path}",
         "https://catbox.moe/user/api.php"],
        capture_output=True, text=True, timeout=30,
    )
    url = result.stdout.strip()
    return url if url.startswith("http") else ""


async def upload_to_tmpfiles(tmp_path: str) -> str:
    """Upload a file to tmpfiles.org and return the direct download URL."""
    result = subprocess.run(
        ["curl", "-s", "-F", f"file=@{tmp_path}", "https://tmpfiles.org/api/v1/upload"],
        capture_output=True, text=True, timeout=30,
    )
    try:
        data = json.loads(result.stdout)
        if data.get("status") == "success":
            # Convert page URL to direct download URL: tmpfiles.org/ID/file → tmpfiles.org/dl/ID/file
            page_url = data["data"]["url"]
            return page_url.replace("tmpfiles.org/", "tmpfiles.org/dl/", 1).replace("http://", "https://")
    except (json.JSONDecodeError, KeyError):
        pass
    return ""


async def upload_image(image_bytes: bytes, filename: str = "image.png") -> str:
    """Upload image bytes to a public host. Tries catbox.moe first, tmpfiles.org as fallback."""
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
        f.write(image_bytes)
        tmp_path = f.name

    try:
        url = await upload_to_catbox(tmp_path)
        if url:
            return url
        url = await upload_to_tmpfiles(tmp_path)
        if url:
            return url
        return ""
    finally:
        Path(tmp_path).unlink(missing_ok=True)


# --- MCP Tools ---

@mcp.tool()
async def generate_panda_image(bedrijfsnaam: str, sector: str = "", website: str = "") -> str:
    """
    Genereer een fotorealistische AI Panda mascotte-afbeelding voor een bedrijf.
    Gebruikt Gemini (met referentie-image) als primaire provider en OpenAI als fallback.

    Args:
        bedrijfsnaam: De naam van het bedrijf.
        sector: De sector van het bedrijf (optioneel, voor sector-specifieke omgeving).
        website: Het domein van het bedrijf (optioneel, voor logo ophalen, bijv. "bol.com").
    """
    fallback_url = f"https://ui-avatars.com/api/?name=AI+Panda&size=400&background=000000&color=ffffff&bold=true&format=png"

    # Fetch company logo
    logo_b64 = _fetch_logo_b64(website)

    # Build prompt
    prompt = _build_panda_prompt(bedrijfsnaam, sector, has_logo=bool(logo_b64))

    # Try Gemini (with reference image + logo)
    try:
        image_bytes = await generate_with_gemini(prompt, logo_b64=logo_b64)
        if image_bytes:
            url = await upload_image(image_bytes, f"panda_{bedrijfsnaam.lower().replace(' ', '_')}.png")
            if url:
                return json.dumps({"success": True, "image_url": url, "provider": "gemini", "bedrijfsnaam": bedrijfsnaam})
    except Exception:
        pass

    # Try OpenAI (prompt-only, no reference images)
    try:
        image_bytes = await generate_with_openai(prompt)
        if image_bytes:
            url = await upload_image(image_bytes, f"panda_{bedrijfsnaam.lower().replace(' ', '_')}.png")
            if url:
                return json.dumps({"success": True, "image_url": url, "provider": "openai", "bedrijfsnaam": bedrijfsnaam})
    except Exception:
        pass

    return json.dumps({
        "success": False,
        "error": "Beide providers (Gemini en OpenAI) hebben gefaald.",
        "fallback_url": fallback_url,
    })


@mcp.tool()
async def generate_custom_image(prompt: str) -> str:
    """
    Genereer een willekeurige afbeelding via Gemini (primair) of OpenAI (fallback).
    Retourneer een publieke URL.

    Args:
        prompt: Beschrijving van de gewenste afbeelding (in het Engels voor beste resultaten).
    """
    # Try Gemini
    try:
        image_bytes = await generate_with_gemini(prompt)
        if image_bytes:
            url = await upload_image(image_bytes)
            if url:
                return json.dumps({"success": True, "image_url": url, "provider": "gemini"})
    except Exception:
        pass

    # Try OpenAI
    try:
        image_bytes = await generate_with_openai(prompt)
        if image_bytes:
            url = await upload_image(image_bytes)
            if url:
                return json.dumps({"success": True, "image_url": url, "provider": "openai"})
    except Exception:
        pass

    return json.dumps({"success": False, "error": "Geen afbeelding gegenereerd (Gemini en OpenAI beide gefaald)."})


@mcp.tool()
async def upload_image_base64(image_base64: str, filename: str = "image.png") -> str:
    """
    Upload base64-encoded image data en retourneer een publieke URL.
    Probeert catbox.moe (primair) en tmpfiles.org (fallback). Server-side upload, geen CORS.

    Args:
        image_base64: Base64-encoded image data (zonder data:... prefix).
        filename: Gewenste bestandsnaam (default: image.png).
    """
    try:
        image_bytes = base64.b64decode(image_base64)
        url = await upload_image(image_bytes, filename)
        if not url:
            return json.dumps({"success": False, "error": "Upload mislukt (catbox.moe en tmpfiles.org beide gefaald)."})
        return json.dumps({"success": True, "image_url": url})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


@mcp.tool()
async def read_team_excel() -> str:
    """
    Lees het AI Panda teambestand (ai-panda-team.xlsx) en retourneer alle teamleden als JSON.
    Zoekt automatisch op de juiste locatie (CLAUDE_PLUGIN_ROOT of find fallback).

    Retourneert: JSON array met objecten: [{naam, functie, foto_url, telefoon, email}, ...]
    Bij fout: JSON object met error en searched_paths.
    """
    searched_paths = []

    # Primary: CLAUDE_PLUGIN_ROOT (most reliable in Cowork)
    plugin_root = os.environ.get("CLAUDE_PLUGIN_ROOT", "")
    if plugin_root:
        candidate = Path(plugin_root) / "data" / "ai-panda-team.xlsx"
        searched_paths.append(str(candidate))
        if candidate.exists():
            return _read_excel(candidate)

    # Fallback: relative to this script (local development)
    script_relative = Path(__file__).parent.parent / "data" / "ai-panda-team.xlsx"
    searched_paths.append(str(script_relative))
    if script_relative.exists():
        return _read_excel(script_relative)

    # Fallback: find with timeout
    try:
        result = subprocess.run(
            ["find", "/sessions", str(Path.home()), "-maxdepth", "3", "-name", "ai-panda-team.xlsx"],
            capture_output=True, text=True, timeout=5,
        )
        for line in result.stdout.strip().splitlines():
            path = Path(line.strip())
            searched_paths.append(str(path))
            if path.exists():
                return _read_excel(path)
    except (subprocess.TimeoutExpired, Exception):
        pass

    return json.dumps({
        "error": "ai-panda-team.xlsx niet gevonden",
        "searched_paths": searched_paths,
    })


def _read_excel(path: Path) -> str:
    """Parse the team Excel file and return JSON."""
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active

        headers = [str(cell or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        team = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                entry = {}
                for i, h in enumerate(headers):
                    entry[h] = str(row[i] or "") if i < len(row) else ""
                team.append({
                    "naam": entry.get("naam", ""),
                    "functie": entry.get("functie", ""),
                    "foto_url": entry.get("foto-url", entry.get("foto_url", "")),
                    "telefoon": entry.get("telefoon", ""),
                    "email": entry.get("e-mail", entry.get("email", "")),
                })
        return json.dumps(team, ensure_ascii=False)
    except Exception as e:
        return json.dumps({"error": f"Excel lezen mislukt: {e}", "path": str(path)})


if __name__ == "__main__":
    mcp.run()
